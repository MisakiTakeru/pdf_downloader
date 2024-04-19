# -*- coding: utf-8 -*-
import openpyxl
from urllib.request import urlretrieve
from concurrent.futures import ThreadPoolExecutor
import validators
import itertools
from time import time
import requests



# The three variables to update. 
# excel_path : the excel file to read from
excel_path = 'C:/Users/KOM/Documents/Uge11/test.xlsx'
# column : which column is downloaded on
column = 42
# save_path : Where shall the files be saved to
save_path = 'C:/Users/KOM/Documents/Uge11/pdfs/'


"""""
A function for making a generator for excel sheets. Uses openpyxl to allow
read and write (which makes it about 30% slower to allow write) to get
the headers and values from which is needed to download and save the file.

column[0] == BRNum aka. filename to save as
column[-5] == Pdf_URL
column[-4] == Report Html Adress (file used if Pdf_URL does not work)
column[-1] == downloaded (extra column created to optimize filtering)
"""""
def workbook_gen(workbook):
    columns = workbook.active.columns
    headers = [str(cell.value) for cell in next(columns)]
    headers = [headers[0],headers[-5], headers[-4],headers[-1]]
    for column in columns:
        r = [column[0], column[-5], column[-4], column[-1]]
        yield dict(zip(headers, (cell.value for cell in r)))

"""
Function for my threads to download a pdf file.
Given a single line from an excel file, it checks whether the Pdf_URL is a
valid URL, and if it is not writes down it cannot be downloaded, but if it is
then tries to download and save it, and if this fails runs the auxilliary 
function download_rha and compares it's result.

Parameters
line : dict -- A dictionary of an excel line using the headers as key
i : integer -- a indicator of the position of the current line.
column : integer -- The column where Downloaded is located.
"""
def download(line, i, column):
    downloaded = line['Downloaded']
    if downloaded  == 'yes' or downloaded == 'impossible':
         return 'skip'
    
    url = line['Pdf_URL']

# Validates the URL, if it is malformed or not a correct URL skips it.
    valid = validators.url(url)
    
    if valid:
        download_to = save_path + line['BRnum'] + '.pdf'
        try:
# Tries to download the URL, and if there is no response from server within 3
# seconds gives up.
            response = requests.get(url, timeout = 3)
            open(download_to, 'wb').write(response.content)
            ws.cell(i+2, column, 'yes')
            val = 'success'
        except Exception as e :
            res2 = download_rha(line, i, column)
            if res2 == 'success':
                ws.cell(i+2, column, 'yes')
                val = res2
            else:
                val = 'failed'
# A check from when I used urllib, to use it's errors to check whether
# it was possible to download the file after it failed.
#                try:
#                    match (e, res2):
#                        case ('Not Found' | 'Forbidden', 'Not Found' | 'Forbidden' | 'invalid'):
                ws.cell(i+2, column, 'impossible')
#                        case n :
#                            print(f'new case {n} found')
#                            return 'new case'
#               except:
#                   print('failure with')
#                   print(e)
#                   return 'error'
    else:
        res2 = download_rha(line, i)
        if res2 == 'success':
            ws.cell(i+2, column, 'yes')
            val = res2
        elif res2 == 'Not Found' or res2 == 'Forbidden' or res2 == 'invalid' or res2 == 'Conflict':
            ws.cell(i+2, column, 'impossible')
            val = 'invalid'
        else:
            ws.cell(i+2, column, 'failed')
            val = 'failed'
    return val

"""
Auxilliary function to check and download the URL's from the Report Html Adress
header.
"""
def download_rha(line, i):
    url = line['Report Html Address']
    
    valid = validators.url(url)
    
    if valid:
        download_to = save_path + line['BRnum'] + '.pdf'
        try:
            response = requests.get(url, timeout = 3)
            open(download_to, 'wb').write(response.content)
#            urlretrieve(url, download_to)
            return 'success'
        except Exception as e:
#            print(e)
            return e
    else:
        return 'invalid'


workbook = openpyxl.load_workbook(excel_path)
ws = workbook.active
gen = workbook_gen(workbook)

print('Starting downloading now')

rows = ws.max_row
start = time()
# column for Downloaded, not sure if I make it automatically, or force the user to change it.
with ThreadPoolExecutor(max_workers =16) as executor:
    test = list(itertools.islice(gen, rows))
    executor.map(download, test, range(rows), column, timeout = 5)

end = time()
print(end - start)

workbook.save(excel_path)


# 1000 downloads single thread takes 0.87 hours totaling all data around 17.5 hours.

# 16 threads 1000 downloads took 307.4015212059021 seconds (5.12 minutes), and will be estimated to take 107.89281058458486 minutes

# Actual time spent with all downloads 4966.36762714386 seconds (1.38 hours) (82.77 minutes)

# 1 hour 22 minutes 48 seconds